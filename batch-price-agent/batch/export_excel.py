"""
Excel export: mátrix munkalap + lapos részletes adatok.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from shared.supplier_registry import get_supplier_name

# --- Stílusok ---
_HEADER_FILL    = PatternFill("solid", fgColor="2F5496")
_SUBHDR_FILL    = PatternFill("solid", fgColor="4472C4")
_BEST_FILL      = PatternFill("solid", fgColor="A9D18E")

_HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
_SUBHDR_FONT    = Font(bold=True, color="FFFFFF", size=9)
_BEST_FONT      = Font(bold=True, color="1A3D1F", size=10)
_BOLD           = Font(bold=True, size=10)
_NORMAL         = Font(size=10)
_SMALL          = Font(size=9, color="888888")

_CENTER         = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT           = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_RIGHT          = Alignment(horizontal="right",  vertical="center", wrap_text=False)

_THIN           = Side(style="thin",   color="D0D0D0")
_MED            = Side(style="medium", color="9999BB")
_BORDER         = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_L       = Border(left=_MED,  right=_THIN, top=_THIN, bottom=_THIN)   # shop bal széle


def _fmt_num(value) -> str:
    """12345.6789 → '12 345,68'  (mértékegység a fejlécben van)"""
    if value is None:
        return ""
    try:
        return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    except Exception:
        return str(value)


def _compute_ranks(item: dict, sups: list[str]) -> dict[str, int]:
    priced = [
        (sid, item["suppliers"][sid]["normalized_price_huf"])
        for sid in sups
        if item["suppliers"].get(sid, {}).get("scrape_status") == "ok"
        and item["suppliers"][sid].get("normalized_price_huf") is not None
    ]
    priced.sort(key=lambda x: x[1])
    return {sid: i + 1 for i, (sid, _) in enumerate(priced[:3])}


def _best_offer(item: dict, sups: list[str]) -> tuple[str | None, dict]:
    priced = [
        (sid, item["suppliers"][sid])
        for sid in sups
        if item["suppliers"].get(sid, {}).get("scrape_status") == "ok"
        and item["suppliers"][sid].get("normalized_price_huf") is not None
    ]
    if not priced:
        return None, {}
    return min(priced, key=lambda entry: entry[1]["normalized_price_huf"])


def _autofit(ws, min_w: int = 8, max_w: int = 40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def generate_excel(run_meta: dict, matrix: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _build_matrix_sheet(wb, run_meta, matrix)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Lap 1: Mátrix (UI-val megegyező elrendezés)
# ──────────────────────────────────────────────────────────────────────────────

def _build_matrix_sheet(wb: Workbook, run_meta: dict, matrix: dict):
    ws = wb.create_sheet("Mátrix")

    sups = matrix["selected_suppliers"]
    sup_names = matrix.get("supplier_names", {})

    # Két termékoszlop + három összesítő oszlop, majd webshoponként 3 oszlop
    # Oszlopszélességek
    FIXED_COLS = 5          # Cikkszám + Terméknév + legjobb ajánlat összesítő
    COLS_PER_SUP = 3        # Egységár + Készlet + Link

    total_cols = FIXED_COLS + len(sups) * COLS_PER_SUP

    # ── Fejléc sor 1 ──────────────────────────────────────────────────────────
    # Gép-Coop cikkszám: A1:A2 merge
    for col_idx, label in enumerate(["Gép-Coop cikkszám", "Terméknév"], 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER
        ws.merge_cells(
            start_row=1, start_column=col_idx,
            end_row=2,   end_column=col_idx
        )

    # Legjobb ajánlat összesítője
    summary_header = ws.cell(row=1, column=3, value="Összesítő")
    summary_header.fill = _HEADER_FILL
    summary_header.font = _HEADER_FONT
    summary_header.alignment = _CENTER
    summary_header.border = _BORDER_L
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)

    for col, label in enumerate(["Egységár (HUF/db.)", "Készlet (db.)", "Webshop"], 3):
        c = ws.cell(row=2, column=col, value=label)
        c.fill = _SUBHDR_FILL
        c.font = _SUBHDR_FONT
        c.alignment = _CENTER
        c.border = _BORDER_L if col == 3 else _BORDER

    for s_idx, sid in enumerate(sups):
        sup_name = sup_names.get(sid) or get_supplier_name(sid) or sid
        start_col = FIXED_COLS + s_idx * COLS_PER_SUP + 1
        end_col   = start_col + COLS_PER_SUP - 1

        # Shop neve – colspan=2 merge
        c = ws.cell(row=1, column=start_col, value=sup_name)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER_L
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1,   end_column=end_col
        )

    # ── Fejléc sor 2: aloszlopok ──────────────────────────────────────────────
    for s_idx in range(len(sups)):
        start_col = FIXED_COLS + s_idx * COLS_PER_SUP + 1
        for sub_idx, sub_label in enumerate(["Egységár (Ft/db)", "Készlet (db.)", "Link"]):
            col = start_col + sub_idx
            c = ws.cell(row=2, column=col, value=sub_label)
            c.fill = _SUBHDR_FILL
            c.font = _SUBHDR_FONT
            c.alignment = _CENTER
            c.border = _BORDER_L if sub_idx == 0 else _BORDER

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

    # ── Adatsorok ─────────────────────────────────────────────────────────────
    for r_idx, item in enumerate(matrix["items"], 3):
        ranks = _compute_ranks(item, sups)
        best_sid, best_data = _best_offer(item, sups)

        # Gép-Coop cikkszám
        c = ws.cell(row=r_idx, column=1, value=item["gepcoop_part_no"])
        c.font = Font(bold=True, size=10, name="Courier New")
        c.alignment = _LEFT
        c.border = _BORDER

        # Terméknév
        c = ws.cell(row=r_idx, column=2, value=item.get("product_name") or "")
        c.font = _NORMAL
        c.alignment = _LEFT
        c.border = _BORDER

        # Legjobb ajánlat összesítő
        summary_price = ws.cell(row=r_idx, column=3)
        summary_stock = ws.cell(row=r_idx, column=4)
        summary_shop = ws.cell(row=r_idx, column=5)
        summary_price.border = _BORDER_L
        summary_stock.border = _BORDER
        summary_shop.border = _BORDER

        if best_sid:
            summary_price.value = float(best_data["normalized_price_huf"])
            summary_price.number_format = '#,##0.00'
            summary_price.alignment = _RIGHT

            stock_qty = best_data.get("stock_quantity")
            summary_stock.value = stock_qty if stock_qty is not None else "—"
            summary_stock.alignment = _RIGHT if stock_qty is not None else _CENTER
            if stock_qty is not None:
                summary_stock.number_format = "#,##0"

            summary_shop.value = sup_names.get(best_sid) or get_supplier_name(best_sid) or best_sid
            summary_shop.alignment = _LEFT
        else:
            for summary_cell in (summary_price, summary_stock, summary_shop):
                summary_cell.value = "—"
                summary_cell.font = _NORMAL
                summary_cell.alignment = _CENTER

        for s_idx, sid in enumerate(sups):
            cell_data = item["suppliers"].get(sid, {})
            price_col  = FIXED_COLS + s_idx * COLS_PER_SUP + 1
            stock_col  = price_col + 1
            link_col   = price_col + 2
            rank       = ranks.get(sid)

            mapping_ok = cell_data.get("mapping_status") == "mapped"
            status     = cell_data.get("scrape_status", "")

            # -- Egységár cella --
            price_cell = ws.cell(row=r_idx, column=price_col)
            price_cell.border = _BORDER_L

            # -- Készlet cella --
            stock_cell = ws.cell(row=r_idx, column=stock_col)
            stock_cell.border = _BORDER

            # -- Link cella --
            link_cell = ws.cell(row=r_idx, column=link_col)
            link_cell.border = _BORDER
            link_cell.alignment = _CENTER

            if not mapping_ok or status in ("skipped", "pending", "no_mapping", ""):
                price_cell.value     = "Hiányzó mapping"
                price_cell.font      = _NORMAL
                price_cell.alignment = _CENTER
                stock_cell.value     = "—"
                stock_cell.font      = _NORMAL
                stock_cell.alignment = _CENTER
                link_cell.value      = ""

            elif status != "ok":
                err_map = {
                    "not_found":    "Termék nincs a webshopban",
                    "not_priced":   "Termék nincs beárazva",
                    "timeout":      "Technikai hiba",
                    "login_failed": "Technikai hiba",
                }
                err_text = err_map.get(status, "Technikai hiba")
                price_cell.value     = err_text
                price_cell.font      = _NORMAL
                price_cell.alignment = _CENTER
                stock_cell.value     = "—"
                stock_cell.font      = _NORMAL
                stock_cell.alignment = _CENTER
                link_cell.value      = ""

            else:
                price_huf = cell_data.get("normalized_price_huf")
                stock_qty = cell_data.get("stock_quantity")
                product_url = cell_data.get("product_url")

                price_cell.value     = _fmt_num(price_huf) if price_huf is not None else "—"
                price_cell.alignment = _RIGHT

                # Kizárólag a legjobb webshop egységár-cellája kap zöld kiemelést.
                if rank == 1:
                    price_cell.fill = _BEST_FILL
                    price_cell.font = _BEST_FONT
                else:
                    price_cell.font = _NORMAL

                if stock_qty is not None:
                    stock_cell.value         = stock_qty
                    stock_cell.number_format = "#,##0"
                    stock_cell.alignment     = _RIGHT
                else:
                    stock_cell.value     = "—"
                    stock_cell.font      = _SMALL
                    stock_cell.alignment = _CENTER

                if product_url:
                    link_cell.value = f'=HYPERLINK("{product_url}","Megnyitás")'
                    link_cell.font  = Font(color="0563C1", underline="single", size=9)
                else:
                    link_cell.value = "—"
                    link_cell.font  = _SMALL

        ws.row_dimensions[r_idx].height = 18

    # ── Oszlopszélességek ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

    for s_idx in range(len(sups)):
        price_col = FIXED_COLS + s_idx * COLS_PER_SUP + 1
        stock_col = price_col + 1
        link_col  = price_col + 2
        ws.column_dimensions[get_column_letter(price_col)].width = 22
        ws.column_dimensions[get_column_letter(stock_col)].width = 13
        ws.column_dimensions[get_column_letter(link_col)].width  = 14

    # ── Meta adatok a lap tetején (külön lap helyett) ─────────────────────────
    # Projekt neve / dátum a lap nevében és egy kezdeti megjegyzés-sorban sem kell,
    # mert a run_meta adatai a Részletes lapon megvannak.


# ──────────────────────────────────────────────────────────────────────────────
# Lap 2: Lapos részletes adatok (gépi feldolgozásra)
# ──────────────────────────────────────────────────────────────────────────────

_STATUS_LABEL = {
    "ok":           "OK",
    "not_found":    "Termék nincs a webshopban",
    "not_priced":   "Termék nincs beárazva",
    "timeout":      "Technikai hiba",
    "login_failed": "Technikai hiba",
    "no_mapping":   "Hiányzó mapping",
    "skipped":      "Hiányzó mapping",
    "pending":      "Hiányzó mapping",
}


def _build_detail_sheet(wb: Workbook, run_meta: dict, matrix: dict):
    ws = wb.create_sheet("Részletes adatok")
    ws.freeze_panes = "A2"

    cols = [
        "Projekt neve", "Futás azonosító", "Lekérdezés dátuma",
        "Gép-Coop cikkszám", "Terméknév",
        "Webshop ID", "Webshop neve", "Beszállítói cikkszám",
        "Nyers ár", "Pénznem", "Mennyiségi egység",
        "Normalizált Ft/db", "Készlet (db.)", "Státusz", "Rang", "Hibaüzenet",
    ]
    for c_idx, label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    project_name = run_meta.get("project_name", "")
    batch_run_id = run_meta.get("batch_run_id", "")
    created_at   = run_meta.get("created_at", "")
    if isinstance(created_at, str) and "T" in created_at:
        try:
            created_at = datetime.fromisoformat(created_at).strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass

    sups = matrix["selected_suppliers"]

    r = 2
    for item in matrix["items"]:
        pn           = item["gepcoop_part_no"]
        product_name = item.get("product_name") or ""
        ranks        = _compute_ranks(item, sups)

        for sid in sups:
            cell_data = item["suppliers"].get(sid, {})
            price_huf = cell_data.get("normalized_price_huf")
            rank      = ranks.get(sid, "")

            row_data = [
                project_name, batch_run_id, created_at,
                pn, product_name,
                sid, get_supplier_name(sid), cell_data.get("supplier_part_no") or "",
                cell_data.get("raw_price") or "",
                cell_data.get("currency") or "",
                cell_data.get("unit") or "",
                price_huf if price_huf is not None else "",
                cell_data.get("stock_quantity") if cell_data.get("stock_quantity") is not None else "",
                _STATUS_LABEL.get(cell_data.get("scrape_status") or "", cell_data.get("scrape_status") or "Hiányzó mapping"),
                rank,
                cell_data.get("error_message") or "",
            ]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font      = _NORMAL
                cell.alignment = _LEFT
                cell.border    = _BORDER

            r += 1

    _autofit(ws)
    ws.row_dimensions[1].height = 22
